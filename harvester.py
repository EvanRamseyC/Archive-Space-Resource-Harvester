# Import the following libraries
import os
import requests
from concurrent.futures import ThreadPoolExecutor
import openpyxl


def scope_of_search(range_tuple):
    answer = input(f'Do you want to renew from a previous file that did not complete. Y/N?\n').lower()
    if answer == 'n':
        start = int(input(f'Where do you want to start between {range_tuple[0]} and {range_tuple[1]}?\n'))
        end = int(input(f'Where do you want to end between {range_tuple[0]} and {range_tuple[1]}?\n'))
        if end > range_tuple[1]:
            print(f'You cannot end at {end} because the highest index is {range_tuple[1]}')
            end = range_tuple[1]
        if start < range_tuple[0]:
            print(f'You cannot start at {start} because the lowest index is {range_tuple[0]}')
            start = range_tuple[0]
        return start, end
    elif answer == 'y':
        workbook, book_name = where_is_file()
        return workbook, book_name
    else:
        exit()


def where_is_file():
    # This function finds the file that you want to write to.
    file_name = input(f'What is the name of the file you want to continue? Include .xlsx\n')
    workbook, file_found = find_file(file_name)
    if file_found is False:
        f_path = input("Please input the directory the file is in. Enter 'exit' to quit.\n")
        if f_path == 'exit':
            exit()
        elif os.path.exists(f_path):
            os.chdir(f_path)
            workbook, file_found = find_file(file_name)
            if file_found is False:
                print(f'Could not find the file {file_name} in {f_path}')
                exit()
    if file_found is True:
        return workbook, file_name


def find_file(file_name):
    try:
        wb = openpyxl.load_workbook(file_name)
        if "ArchiveSpace Data" and "Log" in wb.sheetnames:
            print(f'Found the file {file_name}')
            return wb, True
        else:
            print("The file you entered needs the sheets, 'ArchiveSpace Data' and 'Log' to continue.")
            exit()
    except FileNotFoundError:
        print(f'Could not find the file {file_name}')
        return False, False


def load_file(wb):
    sheet = wb['Log']
    start, end, last_left = sheet['A2'].value, sheet['B2'].value, sheet['B3'].value
    sheet = wb['ArchiveSpace Data']
    row = sheet.max_row + 1
    return last_left, end, row



def resource_title(resource_record):
    resource_dict = {'title': resource_record['title'], 'uri': resource_record['uri']}
    return resource_dict


def reformat_uri(object_uri):
    if "resources" in object_uri:
        substring = object_uri.split("resources/")[-1]
        substring = f'#tree::resources/{substring}'
        return substring
    if "archival_objects" in object_uri:
        substring = object_uri.split("archival_objects/")[-1]
        substring = f'#tree::archival_object_{substring}'
        return substring
    return


def json_writer(current_record, parent, resource_number, base_url, headers):
    date_expression = []
    begin_date = []
    end_date = []
    scopecontents = []
    abstract = []
    bio_hist = []
    new_uri = reformat_uri(current_record['uri'])
    # The uri needs to be changed to FSU's uri format.
    title = current_record.get('title', current_record.get('display_string', 'No title'))
    child_dict = {'title': title, 'parent_resource': parent,
                  'uri': f'https://sandbox.archivesspace.org/staff/resources/{resource_number}{new_uri}'}
    for date_instance in current_record['dates']:
        if 'expression' in date_instance.keys():
            date = f"{date_instance['expression']}"
            date_expression.append(date)
        if 'begin' in date_instance.keys():
            date = f"{date_instance['begin']}"
            begin_date.append(date)
        if 'end' in date_instance.keys():
            date = f"{date_instance['end']}"
            end_date.append(date)
    child_dict['date expression'] = date_expression
    child_dict['begin date'] = begin_date
    child_dict['end date'] = end_date

    if len(current_record['subjects']) > 0:
        subjects = []
        for subject in current_record['subjects']:
            subjects.append(subject['ref'])
        subjects = linked_data_name(subjects, base_url, headers)
        child_dict['subjects'] = subjects
    if len(current_record['notes']) > 0:
        notes = current_record['notes']
        for note in notes:
            if note['type'] == 'bioghist':
                for content in note['subnotes']:
                    bio_hist.append(content['content'])
                child_dict['bio_hist'] = bio_hist
            if note['type'] == 'scopecontent':
                for content in note['subnotes']:
                    scopecontents.append(content['content'])
                child_dict['scope&contents'] = scopecontents
            if note['type'] == 'abstract':
                abstract.append(note['content'])
            child_dict['abstract'] = scopecontents
    if len(current_record['linked_agents']) > 0:
        persons = []
        for person in current_record['linked_agents']:
            persons.append(person['ref'])
        persons = linked_data_name(persons, base_url, headers)
        child_dict['people'] = persons
    return child_dict


# Writes archival objects without children to json
def json_request_maker(no_child_list, parent, resource_number, baseURL, headers):
    dict_list = []
    with ThreadPoolExecutor(max_workers=50) as pool:
        results = list(pool.map(no_children_request, no_child_list, [baseURL] * len(no_child_list),
                                [headers] * len(no_child_list)))
    for json in results:
        data = json_writer(json, parent, resource_number, baseURL, headers)
        if data is not False:
            dict_list.append(data)
    return dict_list


def linked_data_name(subject_uris, baseURL, headers):
    subject_names = []
    for subject_uri in subject_uris:
        subject_request = requests.get(baseURL + subject_uri, headers=headers).json()
        subject_names.append(subject_request['title'])
    print(subject_names)
    return subject_names


# Gets json files for all objects under the resource and the resource
def no_children_request(no_child_uri, baseURL, headers):
    json_request = requests.get(baseURL + no_child_uri, headers=headers).json()
    # print(json_request['uri'])
    return json_request


def uri_adder(previous_uris, uri_to_be_added):
    if type(uri_to_be_added) is str:
        updated_uris = previous_uris.append(uri_to_be_added)
        return updated_uris
    if type(uri_to_be_added) is list:
        for x in range(len(uri_to_be_added)):
            previous_uris.append(uri_to_be_added[x])
        updated_uris = previous_uris
        return updated_uris


def one_level_down(children_uri, baseURL, headers, resource_uri):
    children_tree = arch_obj_tree(children_uri, baseURL, headers, resource_uri)
    children_filter = children_object_filer(children_uri, children_tree)
    final_uris = []
    if len(children_filter[0]) > 0:
        final_uris = children_filter[0]
    if len(children_filter[1]) > 0:
        final_uris.extend(children_filter[1])
        for x in range(len(children_filter[1])):
            uris = one_level_down((children_filter[1][x]), baseURL, headers, resource_uri)
            final_uris.extend(uris)
    # print(f'Final URIs: {final_uris}')
    return final_uris


def direct_children(parent_resource, parent_resource_uri, parent, baseURL, headers, resource_uri):
    children_list = []
    all_objects = [parent_resource_uri]
    parent_resource_number = parent_resource_uri.split("resources/")[-1]
    if len(parent_resource['precomputed_waypoints']) > 0:
        for x in range(len(parent_resource['precomputed_waypoints'][""]["0"])):
            if parent_resource['precomputed_waypoints'][""]["0"][x]['child_count'] > 0:
                children_list.append(parent_resource['precomputed_waypoints'][""]["0"][x]['uri'])
                all_objects.append(parent_resource['precomputed_waypoints'][""]["0"][x]['uri'])
            else:
                all_objects.append(parent_resource['precomputed_waypoints'][""]["0"][x]['uri'])
        for x in range(len(children_list)):
            uris = one_level_down(children_list[x], baseURL, headers, resource_uri)
            added_uris = uri_adder(all_objects, uris)
            all_objects = added_uris
    data = json_request_maker(all_objects, parent, parent_resource_number, baseURL, headers)
    return data


# The request to get the tree to archival objects. This will not work unless you are requesting data from an archival
# object
def arch_obj_tree(uri, baseURL, headers, resource_uri):
    tree = requests.get(baseURL + resource_uri + "/tree/node?node_uri=" + uri, headers=headers).json()
    return tree


# Creates a list of uris with no children and uris with children
def children_object_filer(parent_uri, g_current_record):
    child_list = []
    send_back_up = []
    # This is the way to find the precomputed waypoints of an archival objects. You cannot use this for resources.
    if len(g_current_record['precomputed_waypoints'].keys()) > 0:
        for z in range(len(g_current_record['precomputed_waypoints'][parent_uri]['0'])):
            if g_current_record['precomputed_waypoints'][parent_uri]['0'][z]['child_count'] == 0:
                send_back_up.append((g_current_record['precomputed_waypoints'][parent_uri]['0'][z]['uri']))
            else:
                child_list.append((g_current_record['precomputed_waypoints'][parent_uri]['0'][z]['uri']))
    return send_back_up, child_list


# Gets uri of archival object
def child_request(uri_list, baseURL, headers, resource_uri):
    grandchild_tree = arch_obj_tree(uri_list, baseURL, headers, resource_uri)
    return grandchild_tree


# Goes further into the archival object tree
def heiarchy_delver(child_list, baseURL, headers, resource_uri):
    uri_list = one_level_down(child_list, baseURL, headers, resource_uri)
    return uri_list


def create_workbook(start, end):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'ArchiveSpace Data'
    workbook.create_sheet("Log")
    create_sheet_headers(worksheet)
    name = f'ArchiveSpace Data({start}-{end}).xlsx'
    worksheet = workbook['Log']
    worksheet['A1'] = 'Start Index'
    worksheet['A2'] = start
    worksheet['A3'] = 'Current index'
    worksheet['B1'] = 'End Index'
    worksheet['B2'] = end
    workbook.save(name)
    return workbook, name


def create_sheet_headers(sheet):
    sheet['A1'] = 'Collection Title'
    sheet['B1'] = 'Parent Resource'
    sheet['C1'] = 'URI'
    sheet['D1'] = 'Date Expression'
    sheet['E1'] = 'Begin Date'
    sheet['F1'] = 'End Date'
    sheet['G1'] = 'Subjects'
    sheet['H1'] = 'Names'
    sheet['I1'] = 'Biographical/Historical'
    sheet['J1'] = 'Scope and Contents'
    sheet['K1'] = 'Abstract'
    sheet['L1'] = 'Misc Notes'
    return


def strip_brackets(string):
    if string.startswith("['") and string.endswith("']"):
        string = string[1:-1]
    if string.startswith("'") and string.endswith("'"):
        string = string[1:-1]
    return string


# Creates the workbook the information will be in
def input_data_to_excel(sheet, book, resource_and_objects_dict, row, book_name):
    title, parent, uri, date_expression, begin_date, end_date, subject, names, biographical, scope, abstract, \
        misc_notes = 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
    for info in resource_and_objects_dict:
        sheet.cell(row=row, column=title).value = strip_brackets(str(info['title']))
        sheet.cell(row=row, column=parent).value = strip_brackets(str(info['parent_resource']))
        sheet.cell(row=row, column=uri).value = info['uri']
        if 'date expression' in info.keys() and len(info['date expression']) > 0:
            sheet.cell(row=row, column=date_expression).value = strip_brackets(str(info['date expression']))
        if 'begin date' in info.keys() and len(info['begin date']) > 0:
            sheet.cell(row=row, column=begin_date).value = strip_brackets(str(info['begin date']))
        if 'end date' in info.keys() and len(info['end date']) > 0:
            sheet.cell(row=row, column=end_date).value = strip_brackets(str(info['end date']))
        if 'subjects' in info.keys() and len(info['subjects']) > 0:
            sheet.cell(row=row, column=subject).value = strip_brackets(str(info['subjects']))
        if 'names' in info.keys() and len(info['names']) > 0:
            sheet.cell(row=row, column=names).value = strip_brackets(str(info['names']))
        if 'bio_hist' in info.keys() and len(info['bio_hist']) > 0:
            sheet.cell(row=row, column=biographical).value = strip_brackets(str(info['bio_hist']))
        if 'scope&contents' in info.keys() and len(info['scope&contents']) > 0:
            sheet.cell(row=row, column=scope).value = strip_brackets(str(info['scope&contents']))
        if 'abstract' in info.keys() and len(info['abstract']) > 0:
            sheet.cell(row=row, column=abstract).value = strip_brackets(str(info['abstract']))
        row += 1
    book.save(book_name)
    return row


def update_sheet_log(sheet, idex, book, book_name):
    sheet['B3'] = idex
    book.save(book_name)
    return


def auth():
    # Your authentication. Replace all of this with your institution's information. This is the authentication for
    # ArchiveSpace's public sandbox
    auth_baseURL = 'https://sandbox.archivesspace.org/staff/api'
    user = 'admin'
    password = 'admin'
    authentication = requests.post(f'https://sandbox.archivesspace.org/staff/api/users/{user}'
                                   f'/login?password={password}').json()

    # Do not change any of this.
    session = authentication['session']
    s_headers = {'X-ArchivesSpace-Session': session, 'Content_Type': 'application/json'}
    print('Your session key is: ' + session)
    repos = '2'
    return auth_baseURL, s_headers, repos


def main():
    baseURL, headers, repository = auth()
    resource_id = requests.get(baseURL + f"/repositories/{repository}/resources?all_ids=true", headers=headers).json()
    # Gets range of all resources
    lowest_index = min(range(len(resource_id)))
    highest_index = max(range(len(resource_id)))
    indices_tuple = (lowest_index, highest_index)
    print(indices_tuple)
    print(f'You have {len(resource_id)} resources in this repository '
          f'starting from {indices_tuple[0]} to {indices_tuple[1]}')

    search_range = scope_of_search(indices_tuple)

    # Create workbook
    if type(search_range[0]) == int:
        resource_start = search_range[0]
        resource_end = search_range[1]
        workbook, workbook_name = create_workbook(resource_start, resource_end)
        row_keeper = 2
    else:
        workbook, workbook_name = search_range
        resource_start, resource_end, row_keeper = load_file(workbook)


    # Goes through each resource, looking at all of their children and putting it on an Excel spreadsheet.
    for i in range(resource_start, resource_end + 1):
        resource_uri = f'/repositories/{repository}/resources/{resource_id[i]}'
        endpoint = f'{resource_uri}/tree/root'
        output = requests.get(baseURL + endpoint, headers=headers).json()
        resource = resource_title(output)

        # Finds information for top level data and children
        children = direct_children(output, resource['uri'], resource['title'], baseURL, headers, resource_uri)
        json_dict = children

        # Adds data to Excel spreadsheet
        worksheet = workbook["ArchiveSpace Data"]
        row_keeper = input_data_to_excel(worksheet, workbook, json_dict, row_keeper, workbook_name)
        worksheet = workbook["Log"]
        update_sheet_log(worksheet, i, workbook, workbook_name)

    worksheet['A5'] = 'All resources in this range are done'
    workbook.save(workbook_name)
    print('Done')
    return


if __name__ == '__main__':
    main()
